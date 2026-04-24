import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename)
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['A1']
    # cell = sheet.cell(1, 1)
    # print(cell.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        # print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 5)
        corrected_price_cell.value = corrected_price

    value = Reference(sheet,
                      min_row=2,
                      max_row=sheet.max_row,
                      min_col=5,
                      max_col=5)
    chart = BarChart()
    chart.add_data(value)
    sheet.add_chart(chart, 'F5')

    wb.save(filename)
